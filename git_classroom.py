#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Git Classroom Tool
A Python port of GitService.cs from OutilsProfs, providing interactive GitHub Classroom-like management
using the GitHub CLI (gh) and Git, styled beautifully with the rich library.
"""

import os
import sys
import re
import csv
import json
import tempfile
import shutil
import unicodedata
import subprocess
import concurrent.futures
from datetime import datetime

# Import rich elements for high-fidelity UI
from rich.console import Console
from rich.panel import Panel
from rich.table import Table
from rich.rule import Rule
from rich.prompt import Prompt, Confirm
from rich.progress import Progress, BarColumn, TextColumn, TaskProgressColumn, SpinnerColumn

# Initialize rich console with standard styling
console = Console()

# File extensions that can contain steganographic watermarks
TEXT_EXTENSIONS = {
    ".cs", ".py", ".java", ".js", ".ts", ".jsx", ".tsx",
    ".html", ".css", ".scss", ".txt", ".md", ".rst",
    ".cpp", ".c", ".h", ".hpp", ".php", ".rb", ".go",
    ".vb", ".sql", ".sh", ".bat", ".ps1", ".xml", ".json", ".yaml", ".yml",
}

# Regex to match official repository naming convention
# Pattern: {Session}-{Cours}-{Travail}-{NomFamille}-{Prenom}
# e.g.: H26-4N6-TP1-DUPONT-MARC
REPO_NAME_PATTERN = re.compile(
    r"^([HAE]\d{2})-([A-Z0-9]+)-([A-Z0-9]+)-([A-Z]+)-([A-Z]+)$",
    re.IGNORECASE
)


# ── Slug Helpers ─────────────────────────────────────────────────────────────

def normalize_slug(s: str) -> str:
    """Normalize input string to match our slug/naming convention."""
    if not s:
        return ""
    # Decompose unicode to separate characters from combining accents/diacritics
    normalized = unicodedata.normalize('NFKD', s)
    no_accents = "".join([c for c in normalized if not unicodedata.combining(c)])
    # Keep only alphanumeric characters, uppercase
    cleaned = "".join([c for c in no_accents if c.isalnum()])
    return cleaned.upper()


def build_repo_name(session: str, cours: str, travail: str, nom_famille: str, prenom: str) -> str:
    """Build standardized repo name from metadata."""
    return (
        f"{normalize_slug(session)}-{normalize_slug(cours)}-{normalize_slug(travail)}"
        f"-{normalize_slug(nom_famille)}-{normalize_slug(prenom)}"
    )


# ── Subprocess Helpers ────────────────────────────────────────────────────────

def run_process(cmd: str, args: list, cwd: str = None) -> tuple:
    """Run an external command and return returncode, stdout, and stderr."""
    try:
        res = subprocess.run(
            [cmd] + list(args),
            capture_output=True,
            text=True,
            cwd=cwd,
            encoding='utf-8',
            errors='ignore'
        )
        return res.returncode, res.stdout, res.stderr
    except Exception as ex:
        return -1, "", str(ex)


def run_gh(*args) -> tuple:
    """Run a GitHub CLI (gh) command."""
    return run_process("gh", list(args))


def run_git(work_dir: str, *args) -> tuple:
    """Run a Git command inside a specific working directory."""
    return run_process("git", list(args), cwd=work_dir)


# ── Scope Verification Helpers ────────────────────────────────────────────────

def has_gh_scope(scope: str) -> bool:
    """Check if the authenticated GitHub CLI token has the requested scope."""
    rc, stdout, _ = run_gh("api", "user", "-i")
    if rc != 0:
        return False
    for line in stdout.splitlines():
        if line.lower().startswith("x-oauth-scopes:"):
            return scope.lower() in line.lower()
    return False


def ensure_gh_scope(scope: str) -> bool:
    """Ensure the GitHub token has the required scope. If not, refresh interactively."""
    if has_gh_scope(scope):
        return True

    console.print(f"[yellow]⚠  Le token gh n'a pas le scope [bold]{scope}[/].[/]")
    console.print(f"[dim]Commande requise :[/] [white]gh auth refresh -h github.com -s {scope}[/]")
    console.print()

    if not Confirm.ask("Lancer [bold]gh auth refresh[/] maintenant ?", default=True):
        return False

    try:
        # Run interactively so the user can complete the browser-based authorization
        res = subprocess.run(["gh", "auth", "refresh", "-h", "github.com", "-s", scope])
        if res.returncode == 0 and has_gh_scope(scope):
            return True
    except Exception as ex:
        console.print(f"[red]Erreur lors du refresh : {ex}[/]")
        return False

    console.print(f"[red]Le scope {scope} n'est toujours pas disponible. Action annulée.[/]")
    return False


# ── Steganography (Watermark) ────────────────────────────────────────────────

def encode_watermark(student_id: str) -> str:
    """Encode student identifier into space and tab characters (invisible watermark)."""
    # space = 0, tab = 1
    # Prefix to identify the watermark: 4 pairs of tab-space (\t \t \t \t )
    prefix = "\t \t \t \t "
    data_bytes = student_id.encode('utf-8')
    chars = [prefix]
    for b in data_bytes:
        for i in range(7, -1, -1):
            chars.append('\t' if (b & (1 << i)) != 0 else ' ')
    return "".join(chars)


def decode_watermark(text: str) -> str:
    """Decode student identifier from space and tab characters."""
    prefix = "\t \t \t \t "
    idx = 0
    while True:
        idx = text.find(prefix, idx)
        if idx == -1:
            break
        
        watermark_part = ""
        for char in text[idx + len(prefix):]:
            if char in (' ', '\t'):
                watermark_part += char
            else:
                break
                
        num_bytes = len(watermark_part) // 8
        if num_bytes > 0:
            b_array = bytearray()
            for i in range(num_bytes):
                byte_chars = watermark_part[i*8:(i+1)*8]
                b_val = 0
                for b_idx, c in enumerate(byte_chars):
                    if c == '\t':
                        b_val |= (1 << (7 - b_idx))
                b_array.append(b_val)
            
            try:
                decoded = b_array.decode('utf-8', errors='ignore').strip()
                if decoded:
                    return decoded
            except Exception:
                pass
        
        idx += len(prefix) + len(watermark_part)
    return ""


def is_likely_binary(file_path: str) -> bool:
    """Check if a file is likely binary by checking for null bytes in first 512 bytes."""
    try:
        with open(file_path, 'rb') as f:
            chunk = f.read(512)
            return b'\x00' in chunk
    except Exception:
        return True


def inject_signatures_in_directory(dir_path: str, student_id: str):
    """Recursively inject invisible steganographic signatures into text files."""
    watermark = encode_watermark(student_id)
    for root, _, files in os.walk(dir_path):
        for file in files:
            ext = os.path.splitext(file)[1].lower()
            if ext not in TEXT_EXTENSIONS:
                continue
            full_path = os.path.join(root, file)
            if is_likely_binary(full_path):
                continue
            try:
                # Read original contents
                with open(full_path, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()
                
                # Append watermark at the end of the first line AND at the end of the file
                lines = content.split('\n')
                if len(lines) > 0 and lines[0].strip():
                    if lines[0].endswith('\r'):
                        lines[0] = lines[0][:-1] + watermark + '\r'
                    else:
                        lines[0] = lines[0] + watermark
                else:
                    lines = [watermark] + lines
                
                new_content = '\n'.join(lines)
                
                # Append watermark to the end of the file as well
                if not new_content.endswith(watermark):
                    new_content += "\n" + watermark
                    
                with open(full_path, 'w', encoding='utf-8') as f:
                    f.write(new_content)
            except Exception:
                pass


# ── Student List Loaders ──────────────────────────────────────────────────────

def load_students_interactive() -> list:
    """Manually input students via terminal."""
    students = []
    console.print("[dim]Entrer les étudiants (ligne vide pour terminer).[/]")
    console.print("[dim]Format : NomFamille Prénom [LoginGitHub][/]")
    while True:
        line = input("  > ").strip()
        if not line:
            break
        parts = line.split(None, 2)
        if len(parts) < 2:
            console.print("[yellow]Format attendu : NomFamille Prénom [LoginGitHub][/]")
            continue
        login = parts[2] if len(parts) >= 3 else None
        students.append({"nom": parts[0], "prenom": parts[1], "login": login})
    return students


def load_students_from_csv(file_path: str) -> list:
    """Load and parse students from a CSV file with auto-detected separators."""
    delim = ','
    try:
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            sample = f.read(2048)
            if ';' in sample:
                delim = ';'
    except Exception:
        pass

    students = []
    nom_col, prenom_col, login_col = 0, 1, 2
    header_parsed = False

    try:
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f, delimiter=delim)
            for row in reader:
                if not row or (row[0] and row[0].strip().startswith('#')):
                    continue

                fields = [cell.strip() for cell in row]

                if not header_parsed:
                    header_parsed = True
                    looks_like_header = any(
                        f.lower() in ("nom", "nom de famille", "lastname", "last name",
                                      "prénom", "prenom", "firstname", "first name",
                                      "login", "github", "username")
                        for f in fields
                    )
                    if looks_like_header:
                        for i, h in enumerate(fields):
                            h_lower = h.lower()
                            if h_lower in ("nom", "nom de famille", "lastname", "last name"):
                                nom_col = i
                            elif h_lower in ("prénom", "prenom", "firstname", "first name"):
                                prenom_col = i
                            elif h_lower in ("login", "github", "username"):
                                login_col = i
                        continue

                if len(fields) < 2:
                    continue

                nom = fields[nom_col] if nom_col < len(fields) else ""
                prenom = fields[prenom_col] if prenom_col < len(fields) else ""
                login = fields[login_col] if login_col < len(fields) else None
                if not login:
                    login = None

                if nom and prenom:
                    students.append({"nom": nom, "prenom": prenom, "login": login})
    except Exception as ex:
        console.print(f"[red]Erreur de lecture du fichier CSV : {ex}[/]")
    return students


def load_students_from_excel(file_path: str) -> list:
    """Load students from an Excel (.xlsx) file using openpyxl."""
    try:
        import openpyxl
    except ImportError:
        console.print("[red]Le module 'openpyxl' est requis pour lire les fichiers Excel (.xlsx).[/]")
        console.print("[dim]Veuillez l'installer avec : pip install openpyxl[/]")
        return []

    students = []
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        
        # Read the first row to determine column indexes
        headers = []
        for cell in ws[1]:
            headers.append((cell.value or "").strip().lower())
        
        nom_col = 1
        prenom_col = 2
        login_col = -1
        
        for i, h in enumerate(headers, 1):
            if h in ("nom", "nom de famille", "lastname", "last name"):
                nom_col = i
            elif h in ("prénom", "prenom", "firstname", "first name", "given name"):
                prenom_col = i
            elif h in ("login", "github", "username", "login github"):
                login_col = i
        
        for row in list(ws.iter_rows(min_row=2, values_only=True)):
            if not any(row):
                continue
            nom = str(row[nom_col - 1]).strip() if (nom_col - 1 < len(row) and row[nom_col - 1] is not None) else ""
            prenom = str(row[prenom_col - 1]).strip() if (prenom_col - 1 < len(row) and row[prenom_col - 1] is not None) else ""
            login = str(row[login_col - 1]).strip() if (login_col > 0 and login_col - 1 < len(row) and row[login_col - 1] is not None) else None
            if not login:
                login = None
            if nom and prenom:
                students.append({"nom": nom, "prenom": prenom, "login": login})
    except Exception as ex:
        console.print(f"[red]Erreur de lecture du fichier Excel : {ex}[/]")
    return students


# ── Interactive Helpers ───────────────────────────────────────────────────────

def prompt_selection(title: str, choices: list) -> str:
    """Displays a custom selection menu to the user."""
    console.print(f"\n[bold]{title}[/]")
    for idx, choice in enumerate(choices, 1):
        console.print(f"  [cyan]{idx:>2}[/] · {choice}")
    console.print()
    while True:
        val = Prompt.ask(f"Votre choix (1-{len(choices)})")
        try:
            num = int(val)
            if 1 <= num <= len(choices):
                return choices[num - 1]
        except ValueError:
            pass
        console.print("[red]Choix invalide. Veuillez entrer un nombre valide.[/]")


def prompt_multi_selection(title: str, choices: list) -> list:
    """Prompt user to select multiple items using ranges and comma-separated numbers."""
    console.print(f"\n[bold red]{title}[/]")
    for idx, choice in enumerate(choices, 1):
        console.print(f"  [[cyan]{idx:>2}[/]] {choice}")
    console.print()
    console.print("[dim]Entrez les numéros séparés par des virgules (ex: 1,3,5) ou une plage (ex: 2-5).[/]")
    console.print("[dim]Laissez vide pour ne rien sélectionner.[/]")
    
    while True:
        val = Prompt.ask("Votre sélection", default="").strip()
        if not val:
            return []
        
        selected_indices = set()
        parts = val.split(',')
        valid = True
        for part in parts:
            part = part.strip()
            if not part:
                continue
            if '-' in part:
                subparts = part.split('-')
                if len(subparts) == 2:
                    try:
                        start = int(subparts[0].strip())
                        end = int(subparts[1].strip())
                        if 1 <= start <= end <= len(choices):
                            selected_indices.update(range(start - 1, end))
                        else:
                            valid = False
                    except ValueError:
                        valid = False
                else:
                    valid = False
            else:
                try:
                    num = int(part)
                    if 1 <= num <= len(choices):
                        selected_indices.add(num - 1)
                    else:
                        valid = False
                except ValueError:
                    valid = False
        
        if valid:
            return [choices[i] for i in sorted(list(selected_indices))]
        console.print("[red]Sélection invalide. Veuillez réessayer.[/]")


def prompt_session() -> str:
    """Prompt for CEGEP session (e.g. H26 for Hiver 2026)."""
    saison = prompt_selection("Saison", ["H – Hiver", "A – Automne", "E – Été"])
    code = saison[0]
    while True:
        annee = Prompt.ask("Année (2 chiffres, ex: 26)").strip()
        if len(annee) == 2 and annee.isdigit():
            return f"{code}{annee}"
        console.print("[red]2 chiffres requis (ex: 26)[/]")


def prompt_visibility() -> str:
    """Prompt for GitHub repository visibility."""
    choice = prompt_selection("Visibilité du repo", ["🔒 Privé (recommandé)", "🌐 Public"])
    return "--private" if "🔒" in choice else "--public"


def prompt_and_load_students() -> list:
    """Determine the student source and load them."""
    mode = prompt_selection(
        "Source de la liste d'étudiants",
        ["✏  Saisir interactivement", "📄 Charger un CSV", "📊 Charger un Excel (.xlsx)"]
    )
    if "CSV" in mode:
        path = Prompt.ask("Chemin du fichier CSV").strip().strip('"')
        if not os.path.exists(path):
            console.print("[red]✘ Fichier introuvable.[/]")
            return []
        return load_students_from_csv(path)
    if "Excel" in mode:
        path = Prompt.ask("Chemin du fichier Excel (.xlsx)").strip().strip('"')
        if not os.path.exists(path):
            console.print("[red]✘ Fichier introuvable.[/]")
            return []
        return load_students_from_excel(path)
    return load_students_interactive()


# ── Organization & Repos List Helpers ─────────────────────────────────────────

def fetch_user_orgs() -> list:
    """Fetch GitHub organizations the authenticated user belongs to."""
    rc, stdout, stderr = run_gh("api", "/user/orgs", "--paginate", "-q", ".[].login")
    if rc != 0:
        return []
    return [line.strip() for line in stdout.splitlines() if line.strip()]


def extract_year(prefix: str) -> str:
    """Try to extract a 2-digit school year (20-35) from the prefix."""
    m = re.match(r'^[a-zA-Z]?(\d{2})', prefix)
    if not m:
        return ""
    try:
        yr = int(m.group(1))
        if 20 <= yr <= 35:
            return m.group(1)
    except Exception:
        pass
    return ""


def detect_work_groups(repos: list) -> list:
    """Group repositories based on common hyphen-separated prefixes."""
    if not repos:
        return []
    
    prefix_map = {}
    for r in repos:
        name = r["name"]
        parts = name.split('-')
        for length in range(1, len(parts)):
            prefix = "-".join(parts[:length])
            prefix_lower = prefix.lower()
            if prefix_lower not in prefix_map:
                prefix_map[prefix_lower] = {"prefix": prefix, "repos": []}
            prefix_map[prefix_lower]["repos"].append(r)
            
    assignment = {}
    for r in repos:
        name = r["name"]
        parts = name.split('-')
        best = None
        for length in range(len(parts) - 1, 0, -1):
            candidate = "-".join(parts[:length]).lower()
            if candidate in prefix_map and len(prefix_map[candidate]["repos"]) >= 2:
                best = prefix_map[candidate]["prefix"]
                break
        assignment[name] = best if best else name
        
    groups = {}
    for r in repos:
        name = r["name"]
        key = assignment[name]
        key_lower = key.lower()
        if key_lower not in groups:
            groups[key_lower] = {"prefix": key, "repos": []}
        groups[key_lower]["repos"].append(r)
        
    grouped_works = []
    for g in groups.values():
        prefix = g["prefix"]
        sorted_repos = sorted(g["repos"], key=lambda x: x["name"])
        year = extract_year(prefix)
        grouped_works.append({
            "prefix": prefix,
            "year": year,
            "repos": sorted_repos
        })
        
    def sort_key(g):
        y_val = g["year"]
        y_key = -int(y_val) if y_val else 999999
        return (y_key, g["prefix"].lower())
        
    return sorted(grouped_works, key=sort_key)


# ── ThreadPool Commits Checker ────────────────────────────────────────────────

def check_repo_commits(r: dict, org: str) -> dict:
    """Helper used to inspect if a student has pushed commits to their repo."""
    rc, stdout, _ = run_gh(
        "api", f"repos/{org}/{r['repo']['name']}/commits", "-X", "GET", "-F", "per_page=2"
    )
    if rc == 0:
        try:
            data = json.loads(stdout)
            if isinstance(data, list) and len(data) > 1:
                return r
        except Exception:
            pass
    return None


# ── Git Classroom Application Class ───────────────────────────────────────────

class GitClassroomApp:
    def __init__(self):
        self.org = None

    def get_or_prompt_org(self) -> str:
        """Reuse or prompt for the target GitHub organization."""
        if self.org and Confirm.ask(f"Utiliser l'organisation [cyan]{self.org}[/] ?", default=True):
            return self.org
        self.org = self.pick_org()
        return self.org

    def pick_org(self) -> str:
        """Select org from a fetched list, or type manually."""
        autre_option = "✏  Saisir manuellement..."
        console.print("[dim]Récupération des organisations...[/]")
        orgs = fetch_user_orgs()
        
        if orgs:
            choices = list(sorted(orgs)) + [autre_option]
            picked = prompt_selection("Organisation GitHub", choices)
            if picked != autre_option:
                return picked
        else:
            console.print("[dim](Impossible de récupérer les organisations — saisie manuelle)[/]")
            
        return Prompt.ask("Organisation GitHub (ex: departement-info-cem)").strip()

    def fetch_org_repos(self, org: str) -> list:
        """Fetch all repositories within the target organization."""
        rc, stdout, stderr = run_gh("repo", "list", org, "--json", "name,createdAt", "--limit", "1000")
        if rc != 0:
            console.print(f"[red]✘ gh repo list : {stderr.strip()}[/]")
            return []
        try:
            return json.loads(stdout)
        except Exception:
            console.print("[red]✘ Erreur de lecture du JSON.[/]")
            return []

    def invite_collaborateur(self, org: str, repo_name: str, login: str):
        """Invite a student as a collaborator with push permissions."""
        rc, _, stderr = run_gh(
            "api", f"repos/{org}/{repo_name}/collaborators/{login}",
            "-X", "PUT", "-f", "permission=push"
        )
        if rc == 0:
            console.print(f"  [green]✔ Collaborateur invité : {login}[/]")
        else:
            console.print(f"  [yellow]⚠ Invitation échouée ({login}) : {stderr.strip()}[/]")

    def seed_repo(self, org: str, repo_name: str, template_dir: str, hide_signature: bool, student_id: str):
        """Seed a repo with files, optional local template directory, and/or invisible signatures."""
        tmp_dir = tempfile.mkdtemp(prefix=f"git_{repo_name}_")
        try:
            if template_dir:
                try:
                    for item in os.listdir(template_dir):
                        s = os.path.join(template_dir, item)
                        d = os.path.join(tmp_dir, item)
                        if item == ".git":
                            continue
                        if os.path.isdir(s):
                            shutil.copytree(s, d, symlinks=True, ignore=shutil.ignore_patterns(".git", ".idea", ".vs", "bin", "obj"))
                        else:
                            shutil.copy2(s, d)
                except Exception as ex:
                    console.print(f"[yellow]  ⚠ Erreur de copie du modèle : {ex}[/]")
            
            if hide_signature:
                inject_signatures_in_directory(tmp_dir, student_id)
                
            repo_url = f"https://github.com/{org}/{repo_name}.git"
            run_git(tmp_dir, "init")
            run_git(tmp_dir, "branch", "-M", "main")
            run_git(tmp_dir, "remote", "add", "origin", repo_url)
            run_git(tmp_dir, "add", ".")
            rc, _, stderr = run_git(tmp_dir, "commit", "-m", "Initial commit")
            if rc != 0:
                console.print(f"[yellow]  ⚠ Commit échoué pour {repo_name}: {stderr.strip()}[/]")
                return
            rc, _, stderr = run_git(tmp_dir, "push", "-u", "origin", "main")
            if rc != 0:
                console.print(f"[yellow]  ⚠ Push échoué pour {repo_name}: {stderr.strip()}[/]")
        finally:
            try:
                shutil.rmtree(tmp_dir)
            except Exception:
                pass

    # ── Menu Operations ───────────────────────────────────────────────────────

    def check_tool(self, tool: str):
        """Check if tool is in PATH and output version."""
        rc, stdout, _ = run_process(tool, ["--version"])
        if rc == 0:
            ver = stdout.splitlines()[0].strip() if stdout else ""
            console.print(f"[green]✔ {tool}[/] : {ver}")
        else:
            console.print(f"[red]✘ {tool}[/] : non trouvé dans le PATH")

    def diagnostic(self):
        """Diagnostic to check tool setups and authentication scopes."""
        console.print(Rule("[cyan]Diagnostic Git / GitHub CLI[/]", style="cyan"))
        console.print()
        self.check_tool("git")
        self.check_tool("gh")
        console.print()
        console.print("[cyan]Authentification GitHub :[/]")
        rc, stdout, stderr = run_gh("auth", "status")
        output = stdout if stdout.strip() else stderr
        for line in output.splitlines():
            line_str = line.strip()
            if not line_str:
                continue
            if rc != 0:
                console.print(f"[red]  {line_str}[/]")
            else:
                console.print(f"[green]  {line_str}[/]")

    def tester_creation_repo(self):
        """Test creating a single repository."""
        console.print(Rule("[cyan]Test – Création d'un repo[/]", style="cyan"))
        console.print()
        
        org = self.get_or_prompt_org()
        default_name = f"TEST-{datetime.utcnow().strftime('%Y%m%d%H%M')}"
        
        repo_name = Prompt.ask("Nom du repo de test", default=default_name).strip()
        visibility = prompt_visibility()
        
        collaborateur = Prompt.ask("Login GitHub du collaborateur (Entrée pour ignorer)", default="").strip()
        collaborateur = collaborateur if collaborateur else None
        
        console.print(f"[dim]Création de [cyan]{org}/{repo_name}[/]...[/]")
        rc, _, stderr = run_gh("repo", "create", f"{org}/{repo_name}", visibility)
        if rc == 0:
            console.print(f"[green]✔ Repo créé : https://github.com/{org}/{repo_name}[/]")
            if collaborateur:
                self.invite_collaborateur(org, repo_name, collaborateur)
                
            if Confirm.ask("Supprimer ce repo de test maintenant ?", default=True):
                rc2, _, stderr2 = run_gh("repo", "delete", f"{org}/{repo_name}", "--yes")
                if rc2 == 0:
                    console.print("[green]✔ Repo supprimé.[/]")
                else:
                    console.print(f"[red]✘ Erreur suppression : {stderr2.strip()}[/]")
        else:
            console.print(f"[red]✘ Échec de création : {stderr.strip()}[/]")

    def creer_repos_etudiants(self):
        """Create a collection of repositories for students in the organization."""
        console.print(Rule("[cyan]Créer les repos d'un TP[/]", style="cyan"))
        console.print()
        
        org = self.get_or_prompt_org()
        session = prompt_session()
        cours = Prompt.ask("Code du cours (ex: 4N6)").strip()
        travail = Prompt.ask("Identifiant du travail (ex: TP1)").strip()
        
        visibility = prompt_visibility()
        
        use_template = Confirm.ask("Copier le contenu d'un dossier local ?", default=False)
        template_dir = None
        if use_template:
            while True:
                template_dir = Prompt.ask("Chemin du dossier local contenant le modèle").strip().strip('"')
                if os.path.exists(template_dir) and os.path.isdir(template_dir):
                    break
                console.print(f"[red]✘ Dossier introuvable ou invalide : {template_dir}[/]")
            
        hide_signature = Confirm.ask("Cacher une signature stéganographique dans les fichiers ?", default=False)
        
        students = prompt_and_load_students()
        if not students:
            console.print("[yellow]⚠ Aucun étudiant chargé.[/]")
            return
            
        # Preview Table
        console.print()
        with_login = sum(1 for e in students if e["login"])
        console.print(f"[cyan]{len(students)} repo(s) à créer dans [bold]{org}[/]"
                      f"{f' · {with_login} login(s) GitHub fourni(s)' if with_login > 0 else ''}[/]")
                      
        preview = Table(show_header=False, border_style="dim")
        preview.add_column("Repo")
        preview.add_column("Étudiant")
        preview.add_column("Login")
        for e in students:
            name = build_repo_name(session, cours, travail, e["nom"], e["prenom"])
            preview.add_row(
                name,
                f"{e['nom']} {e['prenom']}",
                f"[dim]{e['login']}[/]" if e["login"] else "[dim]–[/]"
            )
        console.print(preview)
        
        if not Confirm.ask("Confirmer la création ?", default=False):
            console.print("[dim]Annulé.[/]")
            return
            
        ok, failed = 0, 0
        with Progress(
            TextColumn("[dim]{task.description}[/]"),
            BarColumn(),
            TaskProgressColumn(),
            console=console
        ) as progress:
            task = progress.add_task("Création des repos", total=len(students))
            for e in students:
                repo_name = build_repo_name(session, cours, travail, e["nom"], e["prenom"])
                progress.update(task, description=repo_name)
                
                rc, _, stderr = run_gh("repo", "create", f"{org}/{repo_name}", visibility)
                if rc != 0:
                    console.print(f"[red]  ✘ {repo_name} : {stderr.strip()}[/]")
                    failed += 1
                    progress.advance(task, 1)
                    continue
                    
                if use_template or hide_signature:
                    student_id = f"{e['nom']},{e['prenom']}"
                    self.seed_repo(org, repo_name, template_dir, hide_signature, student_id)
                    
                if e["login"]:
                    self.invite_collaborateur(org, repo_name, e["login"])
                    
                ok += 1
                progress.advance(task, 1)
                
        console.print()
        if ok > 0:
            console.print(f"[green]✔ {ok} repo(s) créé(s).[/]")
        if failed > 0:
            console.print(f"[red]✘ {failed} échec(s).[/]")

    def lister_repos(self):
        """List and group all repositories within the target organization."""
        console.print(Rule("[cyan]Lister les repos[/]", style="cyan"))
        console.print()
        
        org = self.get_or_prompt_org()
        console.print("[dim]Récupération...[/]")
        
        repos = self.fetch_org_repos(org)
        if not repos:
            console.print("[yellow]Aucun repo trouvé.[/]")
            return
            
        groups = detect_work_groups(repos)
        
        table = Table(
            title=f"[cyan]Repos de {org} — {len(repos)} repo(s), {len(groups)} groupe(s)[/]",
            show_header=True,
            header_style="bold"
        )
        table.add_column("Groupe / Préfixe")
        table.add_column("Année", justify="center")
        table.add_column("Nb", justify="right")
        table.add_column("Suffixe étudiant")
        table.add_column("Créé le")
        
        for group in groups:
            first = True
            prefix = group["prefix"]
            for r in group["repos"]:
                name = r["name"]
                suffix = name[len(prefix)+1:] if len(name) > len(prefix) + 1 else name
                
                created_at_str = r["createdAt"]
                try:
                    dt = datetime.strptime(created_at_str[:19], "%Y-%m-%dT%H:%M:%S")
                    date_display = dt.strftime("%Y-%m-%d")
                except Exception:
                    date_display = created_at_str
                
                table.add_row(
                    f"[white]{prefix}[/]" if first else "",
                    f"[cyan]20{group['year']}[/]" if first and group["year"] else ("[dim]–[/]" if first else ""),
                    f"[dim]{len(group['repos'])}[/]" if first else "",
                    suffix,
                    date_display
                )
                first = False
            # Empty separator row between groups
            table.add_row("", "", "", "", "")
            
        console.print(table)

    def cloner_repos_tp(self):
        """Clone all repositories belonging to a work group prefix."""
        console.print(Rule("[cyan]Cloner les repos d'un TP[/]", style="cyan"))
        console.print()
        
        org = self.get_or_prompt_org()
        console.print("[dim]Récupération des repos...[/]")
        
        repos = self.fetch_org_repos(org)
        if not repos:
            console.print("[yellow]Aucun repo trouvé.[/]")
            return
            
        groups = detect_work_groups(repos)
        manual_entry = "✏  Saisir un préfixe manuellement..."
        
        choices = []
        for g in groups:
            yr_suffix = f" · 20{g['year']}" if g['year'] else ""
            choices.append(f"{g['prefix']}  ({len(g['repos'])} repos{yr_suffix})")
        choices.append(manual_entry)
        
        picked = prompt_selection("Choisir le groupe / préfixe à cloner :", choices)
        
        to_clone = []
        if picked == manual_entry:
            prefix = Prompt.ask("Préfixe (les repos commençant par ce texte)").strip()
            to_clone = [r for r in repos if r["name"].lower().startswith(prefix.lower())]
            to_clone = sorted(to_clone, key=lambda x: x["name"])
        else:
            prefix = picked.split("  (")[0]
            for g in groups:
                if g["prefix"].lower() == prefix.lower():
                    to_clone = g["repos"]
                    break
                    
        if not to_clone:
            console.print("[yellow]Aucun repo trouvé pour ce préfixe.[/]")
            return
            
        dest_dir = Prompt.ask("Dossier de destination", default=os.getcwd()).strip().strip('"')
        if not os.path.exists(dest_dir):
            os.makedirs(dest_dir)
            
        console.print(f"[cyan]{len(to_clone)} repo(s) → {dest_dir}[/]")
        if not Confirm.ask("Confirmer ?", default=True):
            return
            
        ok, skipped, failed = 0, 0, 0
        with Progress(
            TextColumn("[dim]{task.description}[/]"),
            BarColumn(),
            TaskProgressColumn(),
            console=console
        ) as progress:
            task = progress.add_task("Clonage", total=len(to_clone))
            for r in to_clone:
                name = r["name"]
                progress.update(task, description=name)
                repo_dest_dir = os.path.join(dest_dir, name)
                if os.path.exists(repo_dest_dir):
                    skipped += 1
                    progress.advance(task, 1)
                    continue
                    
                rc, _, stderr = run_gh("repo", "clone", f"{org}/{name}", repo_dest_dir)
                if rc == 0:
                    ok += 1
                else:
                    failed += 1
                    console.print(f"[red]  ✘ {name}: {stderr.strip()}[/]")
                progress.advance(task, 1)
                
        console.print()
        if ok > 0:
            console.print(f"[green]✔ {ok} cloné(s).[/]")
        if skipped > 0:
            console.print(f"[yellow]⏭  {skipped} déjà présent(s) (ignoré(s)).[/]")
        if failed > 0:
            console.print(f"[red]✘ {failed} échec(s).[/]")

    def lister_repos_avec_commits(self):
        """List repositories matching filters that contain active student commits."""
        console.print(Rule("[cyan]Repos avec commits étudiants[/]", style="cyan"))
        console.print()
        
        org = self.get_or_prompt_org()
        session = Prompt.ask("Session (ex: H26) — laisser vide pour toutes", default="").strip()
        cours = Prompt.ask("Cours — laisser vide pour tous", default="").strip()
        travail = Prompt.ask("Travail (ex: TP1) — laisser vide pour tous", default="").strip()
        
        console.print("[dim]Récupération des repos...[/]")
        repos = self.fetch_org_repos(org)
        if not repos:
            console.print("[yellow]Aucun repo trouvé.[/]")
            return
            
        ses_slug = normalize_slug(session)
        cour_slug = normalize_slug(cours)
        trav_slug = normalize_slug(travail)
        
        candidates = []
        for r in repos:
            m = REPO_NAME_PATTERN.match(r["name"])
            if m:
                ses, cour, trav, nom, pren = m.groups()
                if (not ses_slug or ses.upper() == ses_slug) and \
                   (not cour_slug or cour.upper() == cour_slug) and \
                   (not trav_slug or trav.upper() == trav_slug):
                    candidates.append({
                        "session": ses,
                        "cours": cour,
                        "travail": trav,
                        "nom": nom,
                        "prenom": pren,
                        "repo": r
                    })
                    
        if not candidates:
            console.print("[yellow]Aucun repo correspondant trouvé.[/]")
            return
            
        console.print(f"[dim]Vérification de {len(candidates)} repo(s) (max 5 en parallèle)...[/]")
        
        with_commits = []
        with Progress(
            TextColumn("[dim]{task.description}[/]"),
            BarColumn(),
            TaskProgressColumn(),
            SpinnerColumn(),
            console=console
        ) as progress:
            task = progress.add_task("Vérification", total=len(candidates))
            with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
                futures = {executor.submit(check_repo_commits, c, org): c for c in candidates}
                for f in concurrent.futures.as_completed(futures):
                    res = f.result()
                    if res:
                        with_commits.append(res)
                    progress.advance(task, 1)
                    
        console.print()
        if not with_commits:
            console.print("[yellow]Aucun repo avec commit étudiant trouvé.[/]")
            return
            
        sorted_list = sorted(with_commits, key=lambda x: (x["session"], x["cours"], x["travail"], x["nom"]))
        
        table = Table(
            title=f"[cyan]{len(sorted_list)} repo(s) avec commits étudiants[/]",
            show_header=True,
            header_style="bold"
        )
        table.add_column("Session")
        table.add_column("Cours")
        table.add_column("Travail")
        table.add_column("Étudiant")
        
        last_group = None
        for r in sorted_list:
            group = f"{r['session']}-{r['cours']}-{r['travail']}"
            if last_group and group != last_group:
                table.add_row("", "", "", "")
            last_group = group
            table.add_row(r["session"], r["cours"], r["travail"], f"{r['nom']} {r['prenom']}")
            
        console.print(table)

    def supprimer_anciens(self):
        """Find, select and delete repositories older than 1 year."""
        console.print(Rule("[cyan]Supprimer les repos anciens (> 1 an)[/]", style="cyan"))
        console.print()
        
        org = self.get_or_prompt_org()
        console.print("[dim]Récupération...[/]")
        
        repos = self.fetch_org_repos(org)
        if not repos:
            console.print("[green]✔ Aucun repo trouvé dans cette organisation.[/]")
            return
            
        # Parse and sort by oldest first
        repos_sorted = sorted(repos, key=lambda x: x["createdAt"])
        
        choices = []
        for r in repos_sorted:
            created_at_str = r["createdAt"]
            try:
                dt = datetime.strptime(created_at_str[:19], "%Y-%m-%dT%H:%M:%S")
                date_display = dt.strftime("%Y-%m-%d")
            except Exception:
                date_display = created_at_str
            choices.append(f"{r['name']}  ({date_display})")
            
        selected = prompt_multi_selection(
            f"Sélectionner les repos à supprimer ({len(repos)} repos, plus anciens en premier)",
            choices
        )
        if not selected:
            console.print("[dim]Aucun repo sélectionné. Annulé.[/]")
            return
            
        to_delete_names = [s.split(' ')[0] for s in selected]
        
        console.print()
        console.print(f"[red bold]⚠  {len(to_delete_names)} repo(s) seront supprimés définitivement.[/]")
        console.print("[red]Cette action est irréversible.[/]")
        console.print()
        
        confirm = Prompt.ask("[red]Taper [bold]SUPPRIMER[/] pour confirmer[/]")
        if confirm != "SUPPRIMER":
            console.print("[dim]Annulé.[/]")
            return
            
        # Ensure the gh token has the delete_repo scope before attempting any deletion
        if not ensure_gh_scope("delete_repo"):
            return
            
        ok, failed = 0, 0
        with Progress(
            TextColumn("[dim]{task.description}[/]"),
            BarColumn(),
            TaskProgressColumn(),
            console=console
        ) as progress:
            task = progress.add_task("Suppression", total=len(to_delete_names))
            for name in to_delete_names:
                progress.update(task, description=name)
                rc, _, stderr = run_gh("repo", "delete", f"{org}/{name}", "--yes")
                if rc == 0:
                    ok += 1
                else:
                    failed += 1
                    console.print(f"[red]  ✘ {name}: {stderr.strip()}[/]")
                progress.advance(task, 1)
                
        console.print()
        if ok > 0:
            console.print(f"[green]✔ {ok} repo(s) supprimé(s).[/]")
        if failed > 0:
            console.print(f"[red]✘ {failed} échec(s).[/]")

    def detecter_signatures(self):
        """Scan a folder recursively to detect and decode steganographic signatures."""
        console.print(Rule("[cyan]Détecter les signatures stéganographiques[/]", style="cyan"))
        console.print()
        
        path_to_scan = Prompt.ask("Dossier à analyser", default=os.getcwd()).strip().strip('"')
        if not os.path.exists(path_to_scan):
            console.print(f"[red]✘ Dossier introuvable : {path_to_scan}[/]")
            return
            
        console.print("[dim]Analyse en cours...[/]")
        
        findings = []
        for root, _, files in os.walk(path_to_scan):
            for file in files:
                ext = os.path.splitext(file)[1].lower()
                if ext not in TEXT_EXTENSIONS:
                    continue
                full_path = os.path.join(root, file)
                if is_likely_binary(full_path):
                    continue
                try:
                    with open(full_path, 'r', encoding='utf-8', errors='ignore') as f:
                        content = f.read()
                    sig = decode_watermark(content)
                    if sig:
                        rel_path = os.path.relpath(full_path, path_to_scan)
                        findings.append({
                            "file": file,
                            "rel_path": rel_path,
                            "signature": sig,
                            "full_path": full_path
                        })
                except Exception:
                    pass
                    
        if not findings:
            console.print("[yellow]Aucune signature stéganographique détectée dans ce dossier.[/]")
            return
            
        table = Table(
            title=f"[cyan]{len(findings)} signature(s) détectée(s)[/]",
            show_header=True,
            header_style="bold"
        )
        table.add_column("Fichier")
        table.add_column("Signature détectée")
        table.add_column("Chemin relatif")
        table.add_column("Statut / Alerte")
        
        for f in findings:
            sig_parts = [part.strip().lower() for part in f["signature"].split(",")]
            path_lower = f["rel_path"].lower()
            
            match_found = False
            for part in sig_parts:
                normalized_part = normalize_slug(part).lower()
                if normalized_part and (normalized_part in path_lower or part in path_lower):
                    match_found = True
                    break
            
            status = "[green]OK[/]"
            if not match_found and any(x in path_lower for x in ["-", "_"]):
                status = "[bold red]⚠ Discordance / Copie suspectée ![/]"
                
            table.add_row(
                f["file"],
                f"[bold cyan]{f['signature']}[/]",
                f["rel_path"],
                status
            )
            
        console.print(table)

    def run(self):
        """Submenu execution loop."""
        retour = "← Retour"
        while True:
            console.clear()
            console.print(Rule("[cyan]Git Classroom (Python)[/]", style="cyan"))
            if self.org:
                console.print(f"[dim]Organisation active : {self.org}[/]")
            console.print()
            
            choice = prompt_selection(
                "Choisir une action",
                [
                    "🔍 Diagnostic (gh, git, authentification)",
                    "🧪 Tester la création d'un repo",
                    "➕ Créer les repos d'un TP",
                    "📋 Lister les repos",
                    "📥 Cloner les repos d'un TP",
                    "💬 Repos avec commits étudiants",
                    "🗑  Supprimer les repos anciens (> 1 an)",
                    "🕵️  Détecter les signatures stéganographiques",
                    retour
                ]
            )
            
            console.print()
            if choice == retour:
                break
            elif "Diagnostic" in choice:
                self.diagnostic()
            elif "Tester" in choice:
                self.tester_creation_repo()
            elif "Créer" in choice:
                self.creer_repos_etudiants()
            elif "Lister les repos" in choice:
                self.lister_repos()
            elif "Cloner" in choice:
                self.cloner_repos_tp()
            elif "commits" in choice:
                self.lister_repos_avec_commits()
            elif "Supprimer" in choice:
                self.supprimer_anciens()
            elif "signatures" in choice:
                self.detecter_signatures()
                
            console.print("\n[dim]Appuyez sur Entrée pour revenir au menu principal...[/]")
            input()


if __name__ == "__main__":
    try:
        app = GitClassroomApp()
        app.run()
    except KeyboardInterrupt:
        console.print("\n[dim]Fermeture de l'application.[/]")
        sys.exit(0)
